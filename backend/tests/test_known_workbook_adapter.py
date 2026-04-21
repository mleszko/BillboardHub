from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.import_adapters.fingerprint import header_fingerprint_from_column_names
from app.services.import_adapters.known_workbook_v1 import (
    ADAPTER_ID,
    ADAPTER_VERSION,
    load_workbook,
    matches_xlsx,
)


def _build_workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()


def test_known_workbook_adapter_matches_and_loads_with_fingerprint(monkeypatch) -> None:
    headers = [
        "l.p.",
        "Unnamed: 1",
        "lokalizacja",
        "uwagi",
        "uwagi dot. płatności",
        "wynajmujący",
        "czas trwania umowy",
        "koszt na mc",
        "styczeń",
        "luty",
        "status lipiec 2022",
    ]
    rows = [
        [1, "Białystok", "Centrum", "notatka", "przelew", "ACME", "2026-12-31", 1200, 1000, 1000, "ok"],
    ]
    file_bytes = _build_workbook_bytes(headers, rows)
    fingerprint = header_fingerprint_from_column_names(headers)
    monkeypatch.setattr(
        "app.services.import_adapters.known_workbook_v1.HEADER_FINGERPRINTS",
        frozenset({fingerprint}),
    )

    assert matches_xlsx("synthetic.xlsx", file_bytes, "")
    loaded = load_workbook(file_bytes, "")

    assert loaded.parse_options["adapter_id"] == ADAPTER_ID
    assert loaded.parse_options["adapter_version"] == ADAPTER_VERSION
    assert loaded.guessed_by_model == "adapter:known_workbook_v1"
    assert "miasto" in loaded.source_columns
    assert "status lipiec 2022" not in loaded.source_columns
    assert "styczeń" not in loaded.source_columns
    assert loaded.all_rows
    first = loaded.all_rows[0]
    assert first["miasto"] == "Białystok"
    assert first["uwagi"] == "notatka | przelew"
    by_source = {p.source_column_name: p.target_field_name for p in loaded.proposals}
    assert by_source["miasto"] == "city"
    assert by_source["koszt miesięczny netto"] == "monthly_rent_net"
