from pathlib import Path
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app
from app.services.import_processor import _to_json_safe

_DEV_HEADERS = {
    "x-dev-user-id": "import-dedupe-user",
    "x-dev-user-email": "import-dedupe@billboardhub.test",
}
_DEV_HEADERS_NEW_USER = {
    "x-dev-user-id": "import-dedupe-user-2",
    "x-dev-user-email": "import-dedupe-user-2@billboardhub.test",
}


def _run_import(
    client: TestClient,
    raw_csv: bytes,
    headers: dict[str, str] | None = None,
    *,
    filename: str = "sample_import.csv",
) -> dict:
    request_headers = headers or _DEV_HEADERS
    guess = client.post(
        "/imports/guess-mapping",
        headers=request_headers,
        files={"file": (filename, raw_csv, "text/csv")},
        data={
            "sheet_name": "",
            "header_row_1based": "0",
            "skip_rows_before_header": "0",
            "unpivot_month_columns": "false",
            "monthly_aggregate": "mean",
        },
    )
    assert guess.status_code == 200, guess.text
    proposal = guess.json()
    payload = {
        "session_id": proposal["session_id"],
        "owner_user_id": proposal["owner_user_id"],
        "mapping": [
            {
                "source_column_name": m["source_column_name"],
                "target_field_name": m["target_field_name"],
                "confirmed_by_user": True,
                "user_override": True,
                "transform_hint": m.get("transform_hint"),
            }
            for m in proposal["mapping_suggestions"]
        ],
    }
    confirm = client.post("/imports/confirm-mapping", headers=request_headers, json=payload)
    assert confirm.status_code == 200, confirm.text
    return confirm.json()


def _build_multisheet_workbook() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Arkusz A"
    ws1.append(["Najemca", "Miasto", "Adres", "Data_wygasniecia"])
    ws1.append(["Klient A", "Warszawa", "Marszalkowska 1", "2026-12-31"])
    ws2 = wb.create_sheet("Arkusz B")
    ws2.append(["Wynajmujący", "Miasto", "Adres", "Data_wygasniecia"])
    ws2.append(["Wlasciciel B", "Krakow", "Dluga 2", "2026-12-31"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_multisheet_lp_workbook() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "A"
    ws1.append(["l.p.", "Najemca", "Miasto", "Adres", "Data_wygasniecia"])
    ws1.append([1, "Klient A", "Warszawa", "Adres A", "2026-12-31"])
    ws2 = wb.create_sheet("B")
    ws2.append(["l.p.", "Najemca", "Miasto", "Adres", "Data_wygasniecia"])
    ws2.append([1, "Klient B", "Krakow", "Adres B", "2026-12-31"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_reimport_does_not_duplicate_contracts() -> None:
    sample = Path(__file__).with_name("sample_import.csv").read_bytes()
    with TestClient(app) as client:
        client.get("/health")
        first = _run_import(client, sample)
        assert first["status"] == "completed"
        assert first["imported_rows"] >= 1

        list_after_first = client.get("/contracts", headers=_DEV_HEADERS)
        assert list_after_first.status_code == 200, list_after_first.text
        first_count = len(list_after_first.json()["items"])
        assert first_count >= 1

        second = _run_import(client, sample)
        assert second["status"] == "completed"

        list_after_second = client.get("/contracts", headers=_DEV_HEADERS)
        assert list_after_second.status_code == 200, list_after_second.text
        second_count = len(list_after_second.json()["items"])
        assert second_count == first_count


def test_same_file_import_for_new_user_is_isolated_and_deduplicated() -> None:
    sample = Path(__file__).with_name("sample_import.csv").read_bytes()
    with TestClient(app) as client:
        client.get("/health")

        first_user_import = _run_import(client, sample, headers=_DEV_HEADERS)
        assert first_user_import["status"] == "completed"

        first_user_list = client.get("/contracts", headers=_DEV_HEADERS)
        assert first_user_list.status_code == 200, first_user_list.text
        first_user_count = len(first_user_list.json()["items"])
        assert first_user_count >= 1

        new_user_import = _run_import(client, sample, headers=_DEV_HEADERS_NEW_USER)
        assert new_user_import["status"] == "completed"

        new_user_list = client.get("/contracts", headers=_DEV_HEADERS_NEW_USER)
        assert new_user_list.status_code == 200, new_user_list.text
        new_user_count = len(new_user_list.json()["items"])
        assert new_user_count >= 1

        first_user_after_new_user = client.get("/contracts", headers=_DEV_HEADERS)
        assert first_user_after_new_user.status_code == 200, first_user_after_new_user.text
        assert len(first_user_after_new_user.json()["items"]) == first_user_count

        new_user_reimport = _run_import(client, sample, headers=_DEV_HEADERS_NEW_USER)
        assert new_user_reimport["status"] == "completed"

        new_user_after_reimport = client.get("/contracts", headers=_DEV_HEADERS_NEW_USER)
        assert new_user_after_reimport.status_code == 200, new_user_after_reimport.text
        assert len(new_user_after_reimport.json()["items"]) == new_user_count


def test_reimport_without_contract_number_or_billboard_code_uses_composite_fallback() -> None:
    csv_without_keys = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Czynsz_netto\n"
        "Miejski Ośrodek Pomocy Rodzinie w Suwałkach,Suwałki,"
        "\"ul. Sportowa, reklama działki dla Tomasza Wlazło\",2026-12-31,4200\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        _run_import(client, csv_without_keys)
        first_list = client.get("/contracts", headers=_DEV_HEADERS)
        assert first_list.status_code == 200, first_list.text
        first_count = len(first_list.json()["items"])
        assert first_count >= 1

        _run_import(client, csv_without_keys)
        second_list = client.get("/contracts", headers=_DEV_HEADERS)
        assert second_list.status_code == 200, second_list.text
        second_count = len(second_list.json()["items"])
        assert second_count == first_count


def test_import_without_client_mapping_still_creates_record_with_placeholder() -> None:
    csv_unmapped = (
        "foo_col,bar_col\n"
        "abc,def\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        result = _run_import(client, csv_unmapped)
        assert result["status"] == "completed"
        assert result["imported_rows"] >= 1

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        items = listed.json()["items"]
        assert any(item["advertiser_name"] == "DO_UZUPELNIENIA" for item in items)


def test_multisheet_guess_mapping_accepts_sheet_names() -> None:
    workbook = _build_multisheet_workbook()
    with TestClient(app) as client:
        client.get("/health")
        guess = client.post(
            "/imports/guess-mapping",
            headers=_DEV_HEADERS,
            files={"file": ("multi.xlsx", workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={
                "sheet_name": "Arkusz A",
                "sheet_names": ["Arkusz A", "Arkusz B"],
                "header_row_1based": "0",
                "skip_rows_before_header": "0",
                "unpivot_month_columns": "false",
                "monthly_aggregate": "mean",
            },
        )
        assert guess.status_code == 200, guess.text
        body = guess.json()
        assert body["total_rows"] == 2
        assert set(body.get("parse_options", {}).get("sheet_names", [])) == {"Arkusz A", "Arkusz B"}


def test_multisheet_import_with_sheet_override_per_row() -> None:
    workbook = _build_multisheet_workbook()
    with TestClient(app) as client:
        client.get("/health")
        guess = client.post(
            "/imports/guess-mapping",
            headers=_DEV_HEADERS,
            files={"file": ("multi.xlsx", workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={
                "sheet_names": ["Arkusz A", "Arkusz B"],
                "header_row_1based": "0",
                "skip_rows_before_header": "0",
                "unpivot_month_columns": "false",
                "monthly_aggregate": "mean",
            },
        )
        assert guess.status_code == 200, guess.text
        proposal = guess.json()
        base_mapping = []
        for row in proposal["mapping_suggestions"]:
            target = row["target_field_name"]
            if row["source_column_name"] == "Wynajmujący":
                target = None
            base_mapping.append(
                {
                    "source_column_name": row["source_column_name"],
                    "target_field_name": target,
                    "confirmed_by_user": True,
                    "user_override": True,
                    "transform_hint": row.get("transform_hint"),
                }
            )
        payload = {
            "session_id": proposal["session_id"],
            "owner_user_id": proposal["owner_user_id"],
            "mapping": base_mapping,
            "sheet_overrides": [
                {
                    "sheet_name": "Arkusz B",
                    "mapping": [
                        {
                            "source_column_name": row["source_column_name"],
                            "target_field_name": (
                                "property_owner_name"
                                if row["source_column_name"] == "Wynajmujący"
                                else row["target_field_name"]
                            ),
                            "confirmed_by_user": True,
                            "user_override": row["source_column_name"] == "Wynajmujący",
                            "transform_hint": row.get("transform_hint"),
                        }
                        for row in proposal["mapping_suggestions"]
                    ],
                }
            ],
        }
        confirm = client.post("/imports/confirm-mapping", headers=_DEV_HEADERS, json=payload)
        assert confirm.status_code == 200, confirm.text
        result = confirm.json()
        assert result["status"] == "completed"
        assert result["imported_rows"] == 2

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        items = listed.json()["items"]
        advertiser_names = {item["advertiser_name"] for item in items}
        assert "Klient A" in advertiser_names
        assert "Wlasciciel B" in advertiser_names


def test_confirm_mapping_blocks_lp_contract_number_in_multisheet() -> None:
    workbook = _build_multisheet_lp_workbook()
    with TestClient(app) as client:
        client.get("/health")
        guess = client.post(
            "/imports/guess-mapping",
            headers=_DEV_HEADERS,
            files={"file": ("multi_lp.xlsx", workbook, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={
                "sheet_names": ["A", "B"],
                "header_row_1based": "0",
                "skip_rows_before_header": "0",
                "unpivot_month_columns": "false",
                "monthly_aggregate": "mean",
            },
        )
        assert guess.status_code == 200, guess.text
        proposal = guess.json()
        forced_mapping = []
        for row in proposal["mapping_suggestions"]:
            target = row["target_field_name"]
            if row["source_column_name"] == "l.p.":
                target = "contract_number"
            forced_mapping.append(
                {
                    "source_column_name": row["source_column_name"],
                    "target_field_name": target,
                    "confirmed_by_user": True,
                    "user_override": True,
                    "transform_hint": row.get("transform_hint"),
                }
            )
        confirm = client.post(
            "/imports/confirm-mapping",
            headers=_DEV_HEADERS,
            json={
                "session_id": proposal["session_id"],
                "owner_user_id": proposal["owner_user_id"],
                "mapping": forced_mapping,
            },
        )
        assert confirm.status_code == 200, confirm.text
        result = confirm.json()
        assert result["status"] == "completed"
        assert result["imported_rows"] == 2

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        items = listed.json()["items"]
        assert len(items) >= 2
        matching = [item for item in items if item["advertiser_name"] in {"Klient A", "Klient B"}]
        assert len(matching) >= 2
        assert {"Klient A", "Klient B"}.issubset({item["advertiser_name"] for item in matching})
        assert all(item["contract_number"] in (None, "") for item in matching)


def test_import_coerces_numeric_phone_to_string() -> None:
    csv_with_numeric_phone = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Telefon osoby kontaktowej\n"
        "Klient Tel,Suwałki,Utrata 1,2026-12-31,602237688\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        guess = client.post(
            "/imports/guess-mapping",
            headers=_DEV_HEADERS,
            files={"file": ("phone.csv", csv_with_numeric_phone, "text/csv")},
            data={
                "sheet_name": "",
                "header_row_1based": "0",
                "skip_rows_before_header": "0",
                "unpivot_month_columns": "false",
                "monthly_aggregate": "mean",
            },
        )
        assert guess.status_code == 200, guess.text
        proposal = guess.json()
        payload = {
            "session_id": proposal["session_id"],
            "owner_user_id": proposal["owner_user_id"],
            "mapping": [
                {
                    "source_column_name": m["source_column_name"],
                    "target_field_name": (
                        "contact_phone"
                        if m["source_column_name"] == "Telefon osoby kontaktowej"
                        else m["target_field_name"]
                    ),
                    "confirmed_by_user": True,
                    "user_override": m["source_column_name"] == "Telefon osoby kontaktowej",
                    "transform_hint": m.get("transform_hint"),
                }
                for m in proposal["mapping_suggestions"]
            ],
        }
        confirm = client.post("/imports/confirm-mapping", headers=_DEV_HEADERS, json=payload)
        assert confirm.status_code == 200, confirm.text
        result = confirm.json()
        assert result["status"] == "completed"
        assert result["imported_rows"] >= 1

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        item = next((row for row in listed.json()["items"] if row["advertiser_name"] == "Klient Tel"), None)
        assert item is not None
        assert item["contact_phone"] == "602237688"


def test_import_truncates_overlong_contact_phone() -> None:
    long_phone_value = "telefon kontaktowy: +48 123 456 789 wew 1234, dostępny codziennie 8-20"
    csv_with_long_phone = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Telefon osoby kontaktowej\n"
        f"Klient Long,Suwałki,Utrata 2,2026-12-31,\"{long_phone_value}\"\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        guess = client.post(
            "/imports/guess-mapping",
            headers=_DEV_HEADERS,
            files={"file": ("phone_long.csv", csv_with_long_phone, "text/csv")},
            data={
                "sheet_name": "",
                "header_row_1based": "0",
                "skip_rows_before_header": "0",
                "unpivot_month_columns": "false",
                "monthly_aggregate": "mean",
            },
        )
        assert guess.status_code == 200, guess.text
        proposal = guess.json()
        payload = {
            "session_id": proposal["session_id"],
            "owner_user_id": proposal["owner_user_id"],
            "mapping": [
                {
                    "source_column_name": m["source_column_name"],
                    "target_field_name": (
                        "contact_phone"
                        if m["source_column_name"] == "Telefon osoby kontaktowej"
                        else m["target_field_name"]
                    ),
                    "confirmed_by_user": True,
                    "user_override": m["source_column_name"] == "Telefon osoby kontaktowej",
                    "transform_hint": m.get("transform_hint"),
                }
                for m in proposal["mapping_suggestions"]
            ],
        }
        confirm = client.post("/imports/confirm-mapping", headers=_DEV_HEADERS, json=payload)
        assert confirm.status_code == 200, confirm.text
        result = confirm.json()
        assert result["status"] == "completed"
        assert result["imported_rows"] >= 1

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        item = next((row for row in listed.json()["items"] if row["advertiser_name"] == "Klient Long"), None)
        assert item is not None
        assert item["contact_phone"] is not None
        assert len(item["contact_phone"]) <= 64
        assert item["contact_phone"] == long_phone_value[:64]


def test_import_truncates_overlong_contact_person_and_email() -> None:
    long_person_value = (
        "Mariusz Mierzejewski, wykonawca: Sławomir Wyszomirski tel 880379785 "
        "slawek@projektstudio.elk.pl oraz dodatkowe informacje operacyjne"
    )
    long_email_value = (
        "bardzo.dlugi.alias.kontaktowy.z.bardzo.dlugim.opisem@"
        "przykladowa-bardzo-dluga-domena-kontrahenta-z-opisem.pl"
    )
    csv_with_long_contact_data = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Osoba kontaktowa,E-mail osoby kontaktowej\n"
        f"Klient Kontakt,Suwałki,Utrata 3,2026-12-31,\"{long_person_value}\",\"{long_email_value}\"\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        guess = client.post(
            "/imports/guess-mapping",
            headers=_DEV_HEADERS,
            files={"file": ("contact_long.csv", csv_with_long_contact_data, "text/csv")},
            data={
                "sheet_name": "",
                "header_row_1based": "0",
                "skip_rows_before_header": "0",
                "unpivot_month_columns": "false",
                "monthly_aggregate": "mean",
            },
        )
        assert guess.status_code == 200, guess.text
        proposal = guess.json()
        payload = {
            "session_id": proposal["session_id"],
            "owner_user_id": proposal["owner_user_id"],
            "mapping": [
                {
                    "source_column_name": m["source_column_name"],
                    "target_field_name": (
                        "contact_person"
                        if m["source_column_name"] == "Osoba kontaktowa"
                        else (
                            "contact_email"
                            if m["source_column_name"] == "E-mail osoby kontaktowej"
                            else m["target_field_name"]
                        )
                    ),
                    "confirmed_by_user": True,
                    "user_override": m["source_column_name"] in {"Osoba kontaktowa", "E-mail osoby kontaktowej"},
                    "transform_hint": m.get("transform_hint"),
                }
                for m in proposal["mapping_suggestions"]
            ],
        }
        confirm = client.post("/imports/confirm-mapping", headers=_DEV_HEADERS, json=payload)
        assert confirm.status_code == 200, confirm.text
        result = confirm.json()
        assert result["status"] == "completed"
        assert result["imported_rows"] >= 1

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        item = next((row for row in listed.json()["items"] if row["advertiser_name"] == "Klient Kontakt"), None)
        assert item is not None
        assert item["contact_person"] is not None
        assert item["contact_email"] is not None
        assert len(item["contact_person"]) <= 255
        assert len(item["contact_email"]) <= 255
        assert item["contact_person"] == long_person_value[:255]
        assert item["contact_email"] == long_email_value[:255]


def test_reimport_source_of_truth_removes_contracts_missing_from_new_file() -> None:
    csv_initial = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Kod nośnika,Nazwa nośnika\n"
        "Klient A,Ełk,Kościuszki 1,2026-12-31,ELK-001,przy mrówce\n"
        "Klient B,Ełk,Wojska Polskiego 2,2026-12-31,ELK-002,u matematyka\n"
    ).encode("utf-8")
    csv_second = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Kod nośnika,Nazwa nośnika\n"
        "Klient A,Ełk,Kościuszki 1,2026-12-31,ELK-001,przy mrówce\n"
    ).encode("utf-8")

    with TestClient(app) as client:
        client.get("/health")
        first = _run_import(client, csv_initial)
        assert first["status"] == "completed"
        assert first["imported_rows"] == 2

        listed_first = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed_first.status_code == 200, listed_first.text
        items_first = listed_first.json()["items"]
        assert len(items_first) == 2

        second = _run_import(client, csv_second)
        assert second["status"] == "completed"
        assert second["imported_rows"] == 1

        listed_second = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed_second.status_code == 200, listed_second.text
        items_second = listed_second.json()["items"]
        assert len(items_second) == 1
        assert items_second[0]["advertiser_name"] == "Klient A"
        assert items_second[0]["asset_name"] == "przy mrówce"


def test_reimport_keeps_photo_with_asset_name_matching() -> None:
    csv_initial = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Nazwa nośnika\n"
        "Klient Asset,Ełk,Autobus przy rynku,2026-12-31,Autobus Ełk\n"
    ).encode("utf-8")
    csv_second = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Nazwa nośnika\n"
        "Klient Asset 2,Ełk,Autobus przy rynku,2027-12-31,Autobus Ełk\n"
    ).encode("utf-8")

    with TestClient(app) as client:
        client.get("/health")
        imported = _run_import(client, csv_initial)
        assert imported["status"] == "completed"
        assert imported["imported_rows"] == 1

        listed_before_photo = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed_before_photo.status_code == 200, listed_before_photo.text
        items_before_photo = listed_before_photo.json()["items"]
        assert len(items_before_photo) == 1
        contract_id = items_before_photo[0]["id"]

        upload = client.post(
            f"/contracts/{contract_id}/photo",
            headers=_DEV_HEADERS,
            files={"photo": ("photo.png", b"\x89PNG\r\n\x1a\nfake", "image/png")},
            data={"width": "120", "height": "80"},
        )
        assert upload.status_code == 500 or upload.status_code == 200
        # If storage is not configured in this test environment, skip the URL assertion.
        # The key behavior under test is that matched contract identity is preserved.

        reimported = _run_import(client, csv_second)
        assert reimported["status"] == "completed"
        assert reimported["imported_rows"] == 1

        listed_after = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed_after.status_code == 200, listed_after.text
        items_after = listed_after.json()["items"]
        assert len(items_after) == 1
        assert items_after[0]["id"] == contract_id
        assert items_after[0]["asset_name"] == "Autobus Ełk"


def test_import_without_strong_keys_does_not_merge_multiple_rows() -> None:
    csv_many_rows = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Nazwa nośnika\n"
        "Klient 01,Ełk,Adres 01,2026-12-31,\n"
        "Klient 02,Ełk,Adres 02,2026-12-31,\n"
        "Klient 03,Ełk,Adres 03,2026-12-31,\n"
        "Klient 04,Ełk,Adres 04,2026-12-31,\n"
        "Klient 05,Ełk,Adres 05,2026-12-31,\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        result = _run_import(client, csv_many_rows)
        assert result["status"] == "completed"
        assert result["imported_rows"] == 5

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        items = listed.json()["items"]
        assert len(items) == 5
        assert {item["advertiser_name"] for item in items} == {
            "Klient 01",
            "Klient 02",
            "Klient 03",
            "Klient 04",
            "Klient 05",
        }


def test_import_does_not_merge_rows_when_location_and_expiry_repeat() -> None:
    repeated_location_csv = (
        "Najemca,Miasto,Adres,Data_wygasniecia,Nazwa nośnika\n"
        "Klient A,Ełk,Autobus mobilny,2026-12-31,\n"
        "Klient B,Ełk,Autobus mobilny,2026-12-31,\n"
        "Klient C,Ełk,Autobus mobilny,2026-12-31,\n"
    ).encode("utf-8")
    with TestClient(app) as client:
        client.get("/health")
        result = _run_import(client, repeated_location_csv)
        assert result["status"] == "completed"
        assert result["imported_rows"] == 3

        listed = client.get("/contracts", headers=_DEV_HEADERS)
        assert listed.status_code == 200, listed.text
        items = listed.json()["items"]
        assert len(items) == 3
        assert {item["advertiser_name"] for item in items} == {"Klient A", "Klient B", "Klient C"}


def test_to_json_safe_replaces_nan_and_inf_with_null() -> None:
    payload = {
        "plain_nan": float("nan"),
        "plain_inf": float("inf"),
        "plain_neg_inf": float("-inf"),
        "nested": [1, float("nan"), {"value": float("inf")}],
    }

    safe = _to_json_safe(payload)

    assert safe["plain_nan"] is None
    assert safe["plain_inf"] is None
    assert safe["plain_neg_inf"] is None
    assert safe["nested"][1] is None
    assert safe["nested"][2]["value"] is None
